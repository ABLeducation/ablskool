from django.contrib import admin
from users.models import *
from django_admin_listfilter_dropdown.filters import DropdownFilter, RelatedDropdownFilter, ChoiceDropdownFilter
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from django.http import HttpResponse
import openpyxl
from django.utils.translation import gettext_lazy as _
from unfold.admin import ModelAdmin


class studentadmin(ModelAdmin):
    list_display=('first_name', 'school')
    list_filter=(('school',DropdownFilter),('grade',DropdownFilter),('section',DropdownFilter))
    
class teacheradmin(ModelAdmin):
    list_display=('first_name', 'school')
    list_filter=(('school',DropdownFilter),('grade',DropdownFilter))
# Register your models here.

class useradmin(ModelAdmin):
    search_fields=('username',)
    list_filter=(('is_student',DropdownFilter),('is_teacher',DropdownFilter),)
    
    actions = ['activate_users']

    def activate_users(self, request, queryset):
        queryset.update(is_active=True)
    activate_users.short_description = "Activate selected users"
    
admin.site.register(Contact)
# admin.site.register(User,useradmin)
admin.site.register(user_profile_teacher)
admin.site.register(user_profile_student,studentadmin)
admin.site.register(user_profile_principal)
admin.site.register(user_profile_school)
admin.site.register(user_profile_parent)
admin.site.register(Enquiry)
# admin.site.register(UserLoginActivity)
admin.site.register(Attendance)
admin.site.register(AttendanceReport)
admin.site.register(FeedBackStudent)
admin.site.register(NotificationStudent)
admin.site.register(FeedBackSchool)
admin.site.register(NotificationSchool)
admin.site.register(NotificationPrincipal)
admin.site.register(FeedBackPrincipal)
admin.site.register(StudentInnovativeProject)
admin.site.register(School)
admin.site.register(Macroplanner)
admin.site.register(Microplanner)
admin.site.register(AdvocacyVisit)
admin.site.register(TimeTable)
admin.site.register(Inventory)
admin.site.register(Competition)
admin.site.register(InnovationClub)
admin.site.register(GuestSession)
admin.site.register(KreativityShow)
admin.site.register(SchoolContract)
admin.site.register(SchoolGallery)

class UserActivityAdmin(ModelAdmin):
    list_display = ('user', 'login_time', 'logout_time', 'duration')

admin.site.register(UserActivity, UserActivityAdmin)

class CustomUserAdmin(BaseUserAdmin):
    list_display = ('username', 'email', 'first_name', 'last_name', 'is_staff', 'total_time_spent')

    def total_time_spent(self, obj):
        total_time = UserActivity.get_total_time_spent(obj)
        return UserActivity.format_duration(total_time)

    total_time_spent.short_description = 'Total Time Spent'

    search_fields=('username',)
    list_filter=(('is_student',DropdownFilter),('is_teacher',DropdownFilter),)
    
    actions = ['activate_users']

    def activate_users(self, request, queryset):
        queryset.update(is_active=True)
    activate_users.short_description = "Activate selected users"

admin.site.register(User, CustomUserAdmin)
admin.site.register(ObservationSheet)
admin.site.register(CurriculumView)
admin.site.register(UserActivity1)

class SchoolFilter(admin.SimpleListFilter):
    title = _('school')
    parameter_name = 'school'

    def lookups(self, request, model_admin):
        schools = set()
        for activity in UserLoginActivity.objects.all():
            username = activity.login_username
            if username:
                school = self.get_school_from_username(username)
                if school:
                    schools.add(school)
        return [(school, school) for school in schools]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(login_username__in=[
                user.user.username for user in user_profile_student.objects.filter(school__icontains=self.value())
            ])
        return queryset

    def get_school_from_username(self, username):
        try:
            user_profile = user_profile_student.objects.get(user__username=username)
            return user_profile.school
        except user_profile_student.DoesNotExist:
            return None
        
class ClassFilter(admin.SimpleListFilter):
    title = _('class')
    parameter_name = 'class'

    def lookups(self, request, model_admin):
        grades = set()
        for student in user_profile_student.objects.all():
            grade = student.grade
            if grade:
                grades.add(grade)
        return [(grade, grade) for grade in grades]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(login_username__in=[
                user.user.username for user in user_profile_student.objects.filter(grade__icontains=self.value())
            ])
        return queryset

class SectionFilter(admin.SimpleListFilter):
    title = _('section')
    parameter_name = 'section'

    def lookups(self, request, model_admin):
        sections = set()
        for student in user_profile_student.objects.all():
            section = student.section
            if section:
                sections.add(section)
        return [(section, section) for section in sections]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(login_username__in=[
                user.user.username for user in user_profile_student.objects.filter(section__icontains=self.value())
            ])
        return queryset

@admin.register(UserLoginActivity)
class UserLoginActivityAdmin(ModelAdmin):
    list_display = ('login_username', 'get_student_name','get_grade','get_section','login_datetime','login_num')
    list_filter = (SchoolFilter, ClassFilter, SectionFilter)
    actions = ['export_as_excel']

    def get_student_name(self, obj):
        try:
            student_profile = user_profile_student.objects.get(user__username=obj.login_username)
            return f"{student_profile.first_name} {student_profile.last_name}"
        except user_profile_student.DoesNotExist:
            return None

    get_student_name.short_description = 'Student Name'
    
    def get_grade(self, obj):
        try:
            student_profile = user_profile_student.objects.get(user__username=obj.login_username)
            grade=student_profile.grade
            return f"{grade}"
        except user_profile_student.DoesNotExist:
            return None
    get_grade.short_description = 'Grade'
        
    def get_section(self, obj):
        try:
            student_profile = user_profile_student.objects.get(user__username=obj.login_username)
            section=student_profile.section
            return f"{section}"
        except user_profile_student.DoesNotExist:
            return None

    get_section.short_description = 'Section'

    def export_as_excel(self, request, queryset):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=UserLoginActivities.xlsx'
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'User Login Activities'

        # Define the columns
        columns = ['Username','Student Name','Grade','Section','Login Number','Login DateTime']

        # Write the header row
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_title

        # Write data rows
        for row_num, obj in enumerate(queryset, 2):
            row = [
                obj.login_username,
                self.get_student_name(obj),
                obj.login_num,
                obj.login_datetime,
            ]
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)
        return response

    export_as_excel.short_description = "Export Selected as Excel"
